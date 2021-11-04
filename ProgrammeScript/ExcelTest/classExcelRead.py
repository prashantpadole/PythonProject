import openpyxl


class ExcelOperation:

    def __init__(self,path):
        self.path=path


    def readExcel(self):        
        wb_obj= openpyxl.load_workbook(self.path)
        sheet_obj=wb_obj.active
        maxRow= sheet_obj.max_row
        maxColumn= sheet_obj.max_column

        print("MaxRow is ",maxRow," max Column is ",maxColumn)

        for i in range(1,maxRow+1):
            for j in range(1,maxColumn+1):
                cell_obj=sheet_obj.cell(row=i,column=j)
                print("|",cell_obj.value,end="|")
            print("\n")
        wb_obj.save("C:\MSDE\PythonProject\ProgrammeScript\ExcelTest\ExcelOutputExampleRead.xlsx")



