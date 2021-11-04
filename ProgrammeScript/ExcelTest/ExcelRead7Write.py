import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active

sheet_title = sheet.title

print("acive sheet titel: ",sheet_title)

sheet.title =" RSheet1"
sheet_title = sheet.title

print("acive sheet titel: ",sheet_title)

c1=sheet.cell(row=2,column=1)
c1.value="Prashant"

c2=sheet.cell(row=2,column=2)
c2.value="Padole"

c3=sheet['C2']
c3.value="Capgemini"

wb.save("ExcelOutputExample.xlsx")

wb.create_sheet(index = 2 ,title="NewSheet")

wb.save("ExcelOutputExample.xlsx")
##############
path="C:\MSDE\PythonProject\ProgrammeScript\ExcelOutputExampleRead.xlsx"
class ExcelOperation:
    def __init__(self,path):
        self.path=path
    def readExcel(self):        
        wb_obj= openpyxl.load_workbook(self.path)
        sheet_obj=wb_obj.active
        maxRow= sheet_obj.max_row
        maxColumn= sheet_obj.max_column

        print("MaxRow is ",maxRow," max Column is ",maxColumn)

        for i in range(1,maxRow+1):
            for j in range(1,maxColumn+1):
                cell_obj=sheet_obj.cell(row=i,column=j)
                print("|",cell_obj.value,end="|")
            print("\n"
        #wb.save("C:\MSDE\PythonProject\ProgrammeScript\ExcelOutputExampleRead.xlsx")



exl=ExcelOperation(path)
exl.readExcel()
